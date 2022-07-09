# -*- coding: utf-8 -*-
import os, re
import tabula
import openpyxl
import csv
import shutil
import datetime
####################################################from fmstls.py##########
"""
fpyopenxl(wbN, sheetN):
    Excelfile wbN.xlsx　sheet sheetN Open 
    v1.00
    2022/1/5
    @author: jicc
"""
def fpyopenxl(wbN, sheetN):
    """
    Excelfile wbN.xlsx　sheet sheetN Open

    Parameters
    ----------
    wbN : str
        ExcelFile Name   ex.MH_CowHistory.xlsx
    sheetN : str
        sheet name

    Returns
    -------
    None.

    """
    
    #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    sheet = wb[sheetN]
    return [wb, sheet]
"""
fpyopencsv_robj:
    csvfile Open for Reader object
    v1.00
    2022/1/5
    @author: jicc
"""
def fpyopencsv_robj(csvN):
    """
    csvfile Open for Reader object

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    None.

    """
    #import csv
    
    #filename = csvN.split('.')
    #filename = filename[0]  #拡張子を削除したfilename
    
    filename_file = open(csvN)     #csvfile open
    filename_reader = csv.reader(filename_file)       #get Reader object
    
    
    return filename_reader
    
"""
fpyopencsv_rdata:
    csvfile Open for Reader data
    v1.00
    2022/1/5
    @author: jicc
"""
def fpyopencsv_rdata(csvN):
    """
    csvfile Open for Reader data

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    None.

    """
    #import csv
    
    #filename = csvN.split('.')
    #filename = filename[0]  #拡張子を削除したfilename
    
    filename_file = open(csvN)     #csvfile open
    filename_reader = csv.reader(filename_file)       #get Reader object
    filename_data = list(filename_reader)             #list's list
    
    return filename_data
    
"""
fpyopencsv_w:
    csvfile Open for Writer
    v1.00
    2022/1/7
    @author: jicc
"""
def fpyopencsv_w(csvN):
    """
    csvfile Open for Writer

    Parameters
    ----------
    csvN : str
        csvFile Name   ex.MH_??_History.csv

    Returns
    -------
    None.

    """
    #import csv
    
    output_file = open(csvN, 'w', newline='')       #csvfile open
    output_writer = csv.writer(output_file)       #get Reader object
     
    return output_writer

"""
fpygetCell_value: get value from the target Cell
v1.00
2022/2/4

@author: inoue
"""
def fpygetCell_value(sheet, r, col):
    """
    get value from the target Cell on an Excelsheet

    Parameters
    ----------
    sheet : worksheet
        sheetBLV
    r : int
        row
    col : int
        column

    Returns
    -------
    value

    """

    value = sheet.cell(row=r, column=col).value
    return value

"""
fpyinputCell_value: input value to the target Cell
v1.00
2022/2/4

@author: inoue
"""
def fpyinputCell_value(sheet, r, col, vl):
    """
    input value to the target Cell

    Parameters
    ----------
    sheet : worksheet
        sheetBLV
    r : int
        row
    col : int
        column
    vl : type of value
    
    Returns
    -------
    None.

    """

    sheet.cell(row=r, column=col).value = vl 

#fpyNewSheet#
"""
fpyNewSheet : Excelbookに
sheet　'columns'と同じ sheet　'scolN'を作成する。
ｖ1.01
2022/5/3

@author: jicc

"""
def fpyNewSheet(wbN, sheetN, scolN, r):
    """
    Excelbookに sheet 'scolN' r行目の'columns'を1行目に配置した sheet'sheetN'を作成する。
    *sheet 'columns'(列名を記入したシート) を作成しておく
    Parameters
    ----------
    wbN : 　str          
        sheetを作成するワークブック
    sheetN : str　　　　　　シート名:"????" 
        作成するシート
    scolN : str         シート名: "columns"
        参照するシート
	r : int		r行目 作成するcolumn行
    Returns
    -------
    None.

    """
    #import openpyxl
    
    wb = openpyxl.load_workbook(wbN)
    #sheetN = wb[sheetN]
    wb.create_sheet(title=sheetN, index=0)
    sheet = wb[sheetN]
    scol = wb[scolN]
    
    maxcol = scol.max_column #sheet columnの最終列
    
    for i in range(1,maxcol+1):
        sheet.cell(row=r, column=i).value = scol.cell(row=1, column=i).value
    
     
    wb.save(wbN)
    
"""
fpychgSheetTitle      :change ExcelSheet's title
v1.0
2022/3/30

@author: inoue
"""
def fpychgSheetTitle(wbN, sheetN, sheetN1):
    """
    change the sheet's title

    Parameters
    ----------
    wbN : str
        Excelfile to check double data  '??_CowsHistory.xlsx'
    sheetN : str
        元のシート名  : 'KTFarm'
    sheetN1 : str
        変更名      : 'KTFarmorg' 

    Returns
    -------
    None.

    """
    #import chghistory
    wbobj = fpyopenxl(wbN, sheetN)
    wb = wbobj[0]
    sheet = wbobj[1]
    sheet.title = sheetN1
    wb.save(wbN)
 
####################################################from fmstls.py##########
"""
fpypdf_to_csv
    convert ****.pdf to ****.csv file
v1.00
2022/1/1
@author: inoue
"""

#import tabula

def fpypdf_to_csv(filename, Path):
    '''
    filename: string
    csv変換するpdf  filename

    Returns
    -------
    filename.csv

    '''
    filename_pdf = Path + "\\" +  filename + ".pdf"
    filename_csv = Path + "\\" + filename + ".csv"
    tabula.convert_into(filename_pdf, filename_csv, 
                    stream=True , output_format="csv", pages="all")


"""
fpySpdf_in_Dir  -  ディレクトリ内の特定の拡張子を持つファイルを見つけ
                  file名をプリントする
v1.00
2022/1/1
by jicc
"""
def fpySpdf_in_Dir(Ext, Path):
    
    #import os, re
    
    fs = os.listdir(Path) 
    #Pathに指定したフォルダー内の、ファイル名とフォルダー名のリストを返す
    regex_ext = re.compile(Ext)   #Regex(regular expression)オブジェクトを返す
    #print(regex_ext)
    
    for f in fs:
        #print(f)
        mo = regex_ext.search(f) #regx_extにマッチするとmatdhオブジェクトを返す
        if mo: #!=None
            print(f)

######################################################未使用
"""
fpySpdf_in_Dir_to_csv  -  ディレクトリ内の特定の拡張子(.pdf)を持つファイルを見つけ
                  csvfile に変換する
                  変換したpdffileを　フォルダーpdforgに移動する
v1.01
2022/1/11
by jicc

"""
def fpySpdf_in_Dir_to_csv(Ext, Path, bckPath):
    """
    Parameters
    ----------
    Ext : str
        拡張子　　　'\.pdf'　　
    Path : str
        path      '.\\' カレントディレクトリ
    ｂｃｋPath : str
        file移動するフォルダーのpath

    Returns
    -------
    None.

    """
    
    import os, re
    import chghistory
    import shutil
    fs = os.listdir(Path)
    regex_ext = re.compile(Ext)
    #print(regex_ext)
    
    for f in fs:
        #print(f)
        mo = regex_ext.search(f)
        if mo:
            print(f)
            filename = f.split('.')
            filename = filename[0]
            chghistory.fpypdf_to_csv(filename, Path)
            #print(mo.group())
            filename_pdf = filename + '.pdf'
            shutil.move(filename_pdf, bckPath)
            
            
"""
fpyCowHistory
    牛の個体情報.csvから、CowHistory.csv(changehistory's list )を作成する
    'No'を行頭にいどうすることを中止、単純に個体データと異動データを結合するように変更
    個体識別番号　９桁->１０桁
    日付データ　yyyy.mm.dd -> yyyy/mm/dd の処置を追加
ｖ1.02
2022/1/9
@author: inoue
"""

def fpyCowHistory(csvorgN, csvoutN):
    '''
    牛の個体情報.csvから、CowHistory.csv(changehistory's list )を作成する

    Parameters

    ----------
    csvorgN : str
        もととなるcsvファイル        MH_???_yyyymmdd.csv
    csvoutN : str
        作成するcsvファイル　　　　　　MH_???_History.csv　

    Returns
    -------
    None.

    '''
    #import csv
    #import chghistory
    
    mhcow_file = open(csvorgN)     
    ################################################################
    #csvfile open , encoding="utf-8",  "shift-jis"
    #UnicodeDecodeError: 'cp932' ... のエラーのためencoding="utf-8"を追加でもダメ2022/1/11
    #PS> ps_fpyymd_csvtocowshistory_csv_args.py .csv .\ .\csvorg で実施の時のみ。
    #アナコンダ　インタラクティブシェルで行ったら問題なし。
    #今日試したらError出なかった。　なぜ？？　2022/1/12
    #################################################################
    mhcow_reader = csv.reader(mhcow_file)       #get Reader object
    mhcow_data = list(mhcow_reader)             #list's list
    
    cowhistory_header = mhcow_data[0]
    print(cowhistory_header)
    cowhistory_header = cowhistory_header \
        + ['No', '異動内容', '異動年月日', '住所', '氏名または名称'] #見出し行のリスト
     
    output_file = open(csvoutN, 'w', newline='')
    output_writer = csv.writer(output_file)
    output_writer.writerow(cowhistory_header)
    row_max = mhcow_reader.line_num  # =len(mhcow_data) リストの行数
    
    id_info = mhcow_data[1]  	
    #['個体識別番号', '出生の年月日', '雌雄の別', '母牛の個体識別番号', '種別']
    #1行目の絶対データ
    #print(id_info)
    #id_info_ = id_info[0]
    id_info[0] = fpycsvidNo_9to10( id_info[0] ) #idNo
    id_info[1] = fpydate_dottoslash( id_info[1]) #出生年月日
    id_info[3] = fpycsvidNo_9to10( id_info[3] ) #idNo
    print(id_info)
    for row_num in range(5, row_max):
  
        history = mhcow_data[row_num] 	
        #['No', '異動内容', '異動年月日', '住所', '氏名または名称']
        #row_num行目の相対データ
        #print(history)
        history[2] = fpydate_dottoslash( history[2])
        print(history)
        id_info_history = id_info + history #行データを結合
        output_writer.writerow(id_info_history)
    
    output_file.close()
            
"""
fpycsvlisttoxls: 
    csvfileのデータをexcelfileに移行する
    死亡のテーブルを回避する処置を加えた
    ｖ1.01
    2022/1/13　
    @author: jicc
    
"""
def fpycsvlisttoxls(csvN, wbN, sheetN):
    """
    csvfileのデータをexcelfileに移行する

    Parameters
    ----------
    csvN : str
        original csvfile  'MH_???_History.csv'
    wbN : str
        Excelfile to move History data  'MH_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   'MHFarm' 

    Returns
    -------
    None.

    """
    #import chghistory
    #import openpyxl
    
    wbobj = fpyopenxl(wbN, sheetN)   #get Worksheet object
    wb = wbobj[0]
    sheet = wbobj[1]
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wb[sheetN]
    max_row = sheet.max_row                     
    
    csvdata = fpyopencsv_rdata(csvN)     #list's list of the csvfile
    ln = len(csvdata)       #the length of the list csvdata
    ln_ = len(csvdata[0])   #the number of the csvdata's list[0]
    k = max_row
    for i in range(1, ln):
        No = csvdata[i][5]  #死亡のとき、'異動内容「死亡」の'以下のデータを削除のため
        #* "1", "2", "異動内容「死亡」の"　strlength <=2 で振り分け　2022/1/13 ｖ1.01
        No_ln = len(No)      
        while (No_ln<=2):       #* #No==99まで可能
            for j in range(0, ln_+1):
                if j== 0:
                    sheet.cell(row=max_row+i, column=j+1).value = k
                    #print(k)
                    k = k + 1
                else:
                    sheet.cell(row=max_row+i, column=j+1).value = \
                        csvdata[i][j-1]
                        #l = csvdata[i][j-1]
                        #print(l)
            break #*
            
        
    wb.save(wbN)
    
"""
fpycsvidNo_9to10:
    idNo in a csvfile 9figures to 10figures
    v1.00
    2022/1/9
    @author: inoue
    
"""
def fpycsvidNo_9to10( idNo ):
    """
    idNo in a csvfile 9figures to 10figures

    Parameters
    ----------
    idNo : str
        idNo

    Returns
    -------
    None.

    """
    if len(idNo) == 9:
        idNo = '0' + idNo 
    else:
        idNo = idNo 
    
    return idNo 

"""
fpydate_dottoslash:
    date in a csvfile 'yyyy.mm.dd' to 'yyyy/mm/dd'
    v1.02
    2022/7/6
    @author: inoue
    
"""
def fpydate_dottoslash( date ):
    """
    date in a csvfile 'yyyy.mm.dd' to 'yyyy/mm/dd'

    Parameters
    ----------
    date : str
        date

    Returns
    -------
    date  : datetime 

    """
    date = date.split('.')
    date = "/".join(date)   #*
    #date = datetime.datetime.strptime(date, '%Y/%m/%d')
    #date(str)をdatetimeに変換　ｖ1.01　2022/3/1
    #date = date.strftime('%Y/%m/%d')
    #'yyyy/mm/dd'に変換 v1.02 2022/7/6 #*の状態と同じ解消する2022/7/7
    return date
 
"""
fpyymd_csvtoCowsHistory_csv:
    フォルダー内の個体履歴org(csv)をCowsHistory.csvに変更する
    変更後orgcsvfile を　別フォルダー(./csvorg)に移動する
    v1.01
    2022/1/10
    @author: jicc
"""
def fpyymd_csvtoCowsHistory_csv(Ext, Path, bckPath):
    """
    Parameters
    ----------
    Ext : str
        拡張子　　　'\.csv'　　
    Path : str
        path      '.\\' カレントディレクトリ
    bckPath : str
        file移動するフォルダーのpath　'.\\csvorg' 

    Returns
    -------
    None.

    """
    
    #import os, re
    #import shutil
    #import chghistory
    fs = os.listdir(Path)
    regex_ext = re.compile(Ext)
    #print(regex_ext)
        
    for f in fs:
        #print(f)
        mo = regex_ext.search(f)
        if mo:
            print(f)
            f_ = f.split('.')
            csvoutN = f_[0] + 'H.csv'
            csvorgN = f
            fpyCowHistory(csvorgN, csvoutN)
            
            shutil.move(csvorgN, bckPath)
            #csvoriginalfile(csvodgN) を　フォルダーbckPathに移動
     
            
            
"""
fpyHistory_csvto_xlsx:
    フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する
    移動後???H.csvを　別フォルダー(./csvhistory)に移動する
    v1.02
    2022/7/5
    @author: jicc
"""
def fpyHistory_csvto_xlsx(Ext, Path, bckPath, wbN, sheetN):
    """
    フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する
    移動後???H.csvを　別フォルダー(./csvhistory)に移動する
    add try~except v1.02 2022/7/5
    Parameters
    ----------
    Ext : str
        拡張子　　　'\.csv'　　
    Path : str
        path      '.\\' カレントディレクトリ
    ｂｃｋPath : str
        file移動するフォルダーのpath
    wbN : str
        Excelfile to move History data  'MH_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   'MHFarm' 

    Returns
    -------
    None.

    """
    
    #import os, re
    #import shutil
    #import chghistory
    fs = os.listdir(Path)
    regex_ext = re.compile(Ext)
    #print(regex_ext)
        
    for f in fs:
        #print(f)
        mo = regex_ext.search(f)
        if mo:
            print(f)
            fpycsvlisttoxls(f, wbN, sheetN)
            
            try:
                shutil.move(f, bckPath)
            except shutil.Error:
                print( f + ' already exists')
            #csvoriginalfile(csvodgN) を　フォルダーbckPathに移動
            #上書きできないので例外処理


"""
fpystrtodatetime : str'yyyy/mm/dd'をdatetime に変換する

v1.00
2022/3/1

@author: inoue
"""
def fpystrtodatetime( date ):
    """
    str'yyyy/mm/dd'をdatetime に変換する

    Parameters
    ----------
    date : str
       'yyyy/mm/dd'

    Returns
    -------
    date  : datetime

    """
    #import datetime
    date = datetime.datetime.strptime( date, '%Y/%m/%d')
    
    return date

"""
fpyxlstrymdtodatetime : Excel cell 'yyyy/mm/dd'をdatetimeに変換する

v1.01
2022/3/2

@author: inoue
"""
def fpyxlstrymdtodatetime(wbN, sheetN, col):
    """
    Excel cell 'yyyy/mm/dd'をdatetimeに変換する

    Parameters
    ----------
    wbN : str
        書き換えするExcelFile名   :??_CowsHistory.xlsx
    sheetN : str
        書き換えするシート名　　　　：??Farm
    col : int
        書き換えする列

    Returns
    -------
    None.

    """
    #import fmstls 
    #import chghistory
    #import datetime
    
    xl = []
    xl = fpyopenxl(wbN, sheetN)
    wb = xl[0] #workbook
    sheet = xl[1] #worksheet
    
    for i in range(2, sheet.max_row+1):
        
        date = fpygetCell_value(sheet, i, col)
        if type(date) == str: #date = 'str'の場合datetimtに変換1.01
        #if type(date) != datetime.datetime: #これではNoneセルでstopする
            date = fpystrtodatetime( date )
            fpyinputCell_value(sheet, i, col, date)
        else:
            continue
            
        
        
    wb.save(wbN)

#fpyxllist_to_list#########################################################
"""
fpyxllist_to_list: 
    excelfileのリストを　lists'　list にする
    
    ｖ1.00
    2022/3/9
    @author: jicc
    
"""
def fpyxllist_to_list(wbN, sheetN, ncol):
    """
    excelfileのデータをlists'listにする

    Parameters
    ----------
    wbN : str
        Excelfile to move History data  '??_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   '??Farm' 
    ncol :  int
        number of columns
    Returns
    -------
    xllists : lists' list

    """
    #from jiccModule import chghistory
    #import chghistory
    #import openpyxl
    
    wbobj = fpyopenxl(wbN, sheetN)   #get Worksheet object
    #wb = wbobj[0]
    sheet = wbobj[1]
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wb[sheetN]
    max_row = sheet.max_row
    # max_col = sheet.max_col
    #AttributeError: 'Worksheet' object has no attribute 'max_col'
    xllist = []
    xllists = []
    for i in range(2, max_row+1):  #タイトル行は飛ばす
        
        for j in range(1,ncol+1):
            coldata = sheet.cell(row=i, column=j).value
            xllist.append(coldata)
            
        xllists.append(xllist)
        xllist = []    
    return xllists

#fpyaddclm_to_lsts_lst####################################################
"""
fpyaddclm_to_lsts_lst : 
   lists'listに最終カラムを追加する
   
   v1.0
   2022/3/28

@author: inoue
"""
def fpyaddclm_to_lsts_lst(xllists, colv):
    """
    lists'listに最終カラムを追加する

    Parameters
    ----------
    xllists : lists'list
        lists'list from Excelfile
    colv : int str None etc
        
    Returns
    -------
    最終列を追加した　lists'list 

    """
    
    lxll = len(xllists)
    for i in range(0, lxll):
        xllists[i].append(colv)
    return xllists

#fpydelclm_frm_lsts_lst#################################################
"""
fpydelclm_frm_lsts_lst : 
   lists'listのカラムを削除する
   
   v1.0
   2022/3/28

@author: inoue
"""
def fpydelclm_frm_lsts_lst(xllists, col):
    """
    lists'listのカラムを削除する

    Parameters
    ----------
    xllists : lists'list
        lists'list from ExcelFile
    col : int 
    削除する列番号   
    Returns
    -------
    列を削除した　lists'list 

    """
    
    lxll = len(xllists)
    for i in range(0, lxll):
        del xllists[i][col]
    return xllists

#fpyflag_dblrcd_1#######################################################
"""
fpyflag_dblrcd_1 : flag double record 1
   lists'listの重複リストに　1（重複）でチェックを入れる
   
   v1.01
   2022/4/3

@author: inoue
"""
def fpyflag_dblrcd_1(xllists):
    """
    lists'listの重複リストに　1（重複）でチェックを入れる

    Parameters
    ----------
    xllists : lists'list
        lists'list from Excelfile

    Returns
    -------
    重複リストに "1"を追加した　lists'list 

    """
    
    lxll = len(xllists)
    #xldblrows = []
    for i in range(0, lxll):
        #print(xllists[i])
        #k=0
        
        for j in range(0, i+1):
            #print(xllists[j])
            if j!= i:
                if xllists[i][1:5] == xllists[j][1:5] and xllists[i][7:10] == xllists[j][7:10]:
                #LinNo と No 以外が一致したら v1.01
                    xllists[i][11] = 1
                else:
                    continue
            else:
                continue
            
    return xllists 

#fpydel_dblrcd##############################################################
"""
fpydel_dblrcd : delete double record
   lists'listの重複リストの一つを削除する
   
   v1.0
   2022/3/29

@author: inoue
"""
def fpydel_dblrcd(xllists, colv):
    """
    lists'listの重複リストの一つを削除する

    Parameters
    ----------
    xllists : lists'list
        lists'list from Excelfile
    colv : int
        0, 1

    Returns
    -------
    重複リストに "1"を追加した　lists'list 

    """
    
    lxll = len(xllists)
    xllists_ = []
    for i in range(0, lxll):
        if xllists[i][11] == colv:
            xllists_.append(xllists[i])
        else:
            continue

    return xllists_ 

#fpylisttoxls############################################################
"""
fpylisttoxls: 
    listのデータをexcelfileに移行する
    ｖ1.01
    2022/4/3
    @author: jicc
    
"""
def fpylisttoxls(xllist, wbN, sheetN):
    """
    listのデータをexcelfileに移行する

    Parameters
    ----------
    xllist : str
        original csvfile  'MH_???_History.csv'
    wbN : str
        Excelfile to move History data  'MH_CowsHistory.xlsx'
    sheetN : str
        sheet name to add data   'MHFarm' 

    Returns
    -------
    None.

    """
    #from jiccModule import chghistory
    import chghistory
    #import openpyxl
    
    wbobj = chghistory.fpyopenxl(wbN, sheetN)   #get Worksheet object
    wb = wbobj[0]
    sheet = wbobj[1]
    #wb = openpyxl.load_workbook(wbN)
    #sheet = wb[sheetN]
    #max_row = sheet.max_row                     
    
    ln = len(xllist)
           #the length of the list xllist
    if ln > 0: #リストに要素がない場合を排除 v1.01 2022/4/3
        ln_ = len(xllist[0])   #the number of the xllist's list[0]
        for i in range(0, ln):
            for j in range(0, ln_):
                sheet.cell(row=i+2, column=j+1).value = xllist[i][j]
        #return 1
    #else:
        #return 0 #何もしないと以下にindent errorが出る2022/4/3
 
    wb.save(wbN)        

#fpychk_drecords#########################################################
"""
fpychk_drecords   :check doublue records
    重複データを別シートに抜き出す
v1.0
2022/3/30

@author: inoue
"""
def fpychk_drecords(wbN, sheetN):
    """
    check doublue records
    重複データを別シートに抜き出す
    Parameters
    ----------
    wbN : str
        Excelfile to check double data  '??_CowsHistory.xlsx'
    sheetN : str
        sheet name to check double data   '??Farm'

    Returns
    -------
    None.

    """
    #import chghistory
    #wbobj = chghistory.fpyopenxl(wbN, sheetN)
    #wb = wbobj[0]
    #sheet = wbobj[1]
    
    #excelfileのデータをlists'listにする
    xllists = fpyxllist_to_list(wbN,sheetN, 11)
    #value"0"のカラムflagをすべてのリストに追加する
    xllists_0 = fpyaddclm_to_lsts_lst(xllists, 0)
    #重複データのflagを0->1に変更する
    xllists_01 = fpyflag_dblrcd_1(xllists_0)
    #重複データのないlist
    xllists0 = fpydel_dblrcd(xllists_01, 0)
    #重複していたデータのリスト
    xllists1 = fpydel_dblrcd(xllists_01, 1)
   
    xllists0 = fpydelclm_frm_lsts_lst(xllists0, 11) 
    #col 'flag'の削除
    xllists1 = fpydelclm_frm_lsts_lst(xllists1, 11) 
    #col 'flag'の削除
    
    
    #シート名の変更
    fpychgSheetTitle(wbN, sheetN, sheetN + 'org')
    #振り分け用のシート　KTFarm　と　KTFarmout　を作成する。
    fpyNewSheet(wbN, sheetN, 'columns')
    fpyNewSheet(wbN, sheetN + 'out', 'columns')
    #データを振り分ける
    fpylisttoxls( xllists0, wbN, sheetN)
    fpylisttoxls( xllists1, wbN, sheetN + 'out') 

           
######################################################################
"""
fpychghistoryReference:         reference of chbhistory's functions
ｖ1.1
2022/4/2
@author: jicc
"""
def fpychghistoryReference():
    
    print('-----chghistoryReference ---------------------------------------------------------v1.03------')
    print('**fpyopenxl(wbN, sheetN)')
    print('Excelfile wbN.xlsx　sheet sheetN Open ')
    print('.............................................................................................')
    print('**fpyopencsv_robj(csvN)')
    print('csvfile Open for Reader object')
    print('.............................................................................................')
    print('**fpyopencsv_rdata(csvN)')
    print('csvfile Open for Reader data')
    print('.............................................................................................')
    print('**fpyopencsv_w(csvN)')
    print('csvfile Open for Writer')
    print('.............................................................................................')
    print('**fpypdf_to_csv(filename, Path)')
    print('convert ****.pdf to ****.csv file')
    print('.............................................................................................')
    print('**fpySpdf_in_Dir_to_csv(Ext, Path)')
    print('ディレクトリ内の特定の拡張子(.pdf)を持つファイルを見つけcsvfile に変換する')
    print('.............................................................................................')
    print('**fpyCowHistory(csvorgN, csvoutN)')
    print('牛の個体情報.csvから、CowHistory.csv(changehistory\'s list )を作成する')
    print('.............................................................................................')
    print('**fpycsvlisttoxls(csvN, wbN, sheetN)')
    print('csvfileのデータをexcelfileに移行する')
    print('.............................................................................................')
    print('**fpycsvidNo_9to10( idNo )')
    print('idNo in a csvfile 9figures to 10figures')
    print('.............................................................................................')
    print('**fpydate_dottoslash( date )')
    print('date in a csvfile \'yyyy.mm.dd\' to \'yyyy/mm/dd\'')
    print('.............................................................................................')
    print('**fpyymd_csvtoCowsHistory_csv(Ext, Path, bckPath)')
    print('フォルダー内の個体履歴org(csv)をCowsHistory.csvに変更する')
    print('.............................................................................................')
    print('**fpyHistory_csvto_xlsx(Ext, Path, bckPath, wbN, sheetN)')
    print('フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する')
    print('.............................................................................................')
    print('**fpystrtodatetime( date )')
    print('str\'yyyy/mm/dd\'をdatetime に変換する')
    print('.............................................................................................')
    print('**fpyxlstrymdtodatetime(wbN, sheetN, col)')
    print('Excel cell \'yyyy/mm/dd\'をdatetimeに変換する')
    print('.............................................................................................')
    print('**fpyxllist_to_list(wbN, sheetN, ncol)')
    print('excelfileのリストを　lists\'　list にする')
    print('.............................................................................................')
    print('**fpyaddclm_to_lsts_lst(xllists, colv)')
    print('lists\'listに最終カラムを追加する')
    print('.............................................................................................')
    print('**fpydelclm_frm_lsts_lst(xllists, col)')
    print('lists\'listのカラムを削除する')
    print('.............................................................................................')
    print('**fpyflag_dblrcd_1(xllists)')
    print('lists\'listの重複リストに　1（重複）でチェックを入れる')
    print('.............................................................................................')
    print('**fpydel_dblrcd(xllists, colv)')
    print('lists\'listの重複リストの一つを削除する')
    print('.............................................................................................')
    print('**fpylisttoxls(xllist, wbN, sheetN)')
    print('listのデータをexcelfileに移行する')
    print('....................................................................................')
    print('**fpychgSheetTitle(wbN, sheetN, sheetN1)')
    print('change ExcelSheet\'s title')
    print('....................................................................................')
    print('**fpychk_drecords(wbN, sheetN)')
    print('check doublue records 重複データを別シートに抜き出す')
    print('--------------------------------------------------------------------2022/4/2 by jicc---------')
    
    
"""
fpyCowsHistoryManual:                        マニュアル
ｖ1.0
2022/1/11
@author: jicc
"""
def fpyCowsHistoryManual():
    
    print('-----CowsHistoryManual---------------------------------------------------------v1.01-------')
    print('1.ディレクトリ内(..//CowsHistory)の特定の拡張子(.pdf)を持つファイルを見つけcsvfile に変換する')
    print('	MH_???_yyyymmdd.pdf -> ****.csv テーブル部分のデータ抽出')
    print('	MH_???_yyyymmdd.pdf -> ".\\pdforg\\"へ移行')
    print('   PS> python ps_fpyspdfindirtocsv_args.py Ext Path bckPath')
    print(' Ext: \.pdf, Path: .\\(カレントディレクトリ), bckPath: .\\pdforg')
    print(' ')
    print('2.フォルダー内の個体履歴org(csv)をCowsHistory.csvに変更する')
    print('変更後orgcsvfile を　別フォルダー(./csvorg)に移動する')
    print('	MH_???_yyyymmdd.csv -> ".\\csvorg\\"へ移行')
    print('   PS> ps_fpyymd_csvtocowshistory_csv_args.py Ext Path bckPath')
    print(' Ext: \.csv, Path: .\\(カレントディレクトリ), bckPath: .\\csvorg')
    print(' ')
    print('3.フォルダー内の???H.csv)を???CowsHistory.xlsxに移動する')
    print('移動後???H.csvを　別フォルダー(./csvhistory)に移動する')
    print('	MH_???_yyyymmddH.csv -> ".\\csvhistory\\"へ移行')
    print('   PS> ps_fpyhistory_csvto_xlsx_args.py Ext Path bckPath wbN sheetN')
    print(' Ext: \.csv, Path: .\\(カレントディレクトリ), bckPath: .\\csvhistory')
    print(' wbN: ..\\KT_CowsHistory.xlsx, sheetN:KTFarm')
    print('4.??_CowsHistory.xlsx\/??Farm の　str\"yyyy\/mm\/dd\"を')
    print('datetimeに変換する')
    print('   PS> ps_fpyxlstrymdtodatetime_args.py wbN sheetN　col')
    print(' wbN: ..\\KT_CowsHistory.xlsx, sheetN:KTFarm, col: 3 and 9')
    print('---------------------------------------------------------------2022/3/6 by jicc---------')
    
"""
fpyCowsHistoryTools:                        マニュアル
ｖ1.0
2022/1/11
@author: jicc
"""
def fpyCowsHistoryTools():
    
    print('-----CowsHistoryTools---------------------------------------------------------v1.00-------')
    print('#fpychk_drecords(wbN, sheetN)')
    print('   PS> ps_fpychk_drecords_args.py')
    print(' wbN: ..\\KT_CowsHistory.xlsx, sheetN:KTFarm')
    print('---------------------------------------------------------------2022/4/3 by jicc---------')    
